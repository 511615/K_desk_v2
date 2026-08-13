from __future__ import annotations

import re
from collections.abc import Iterable
from io import BytesIO

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

TITLE_FILL = PatternFill("solid", fgColor="17365D")
SECTION_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
LABEL_FILL = PatternFill("solid", fgColor="EAF2F8")
NOTE_FILL = PatternFill("solid", fgColor="FFF2CC")
POSITIVE_FILL = PatternFill("solid", fgColor="E2F0D9")
NEGATIVE_FILL = PatternFill("solid", fgColor="FCE4D6")
WHITE_FONT = Font(name="Microsoft YaHei", color="FFFFFF", bold=True)
HEADER_FONT = Font(name="Microsoft YaHei", color="17365D", bold=True)
BODY_FONT = Font(name="Microsoft YaHei", color="1F1F1F", size=10)
MUTED_FONT = Font(name="Microsoft YaHei", color="666666", size=9)
THIN_GRAY = Side(style="thin", color="D9E2F3")
MONEY_FORMAT = '#,##0.00;[Red]-#,##0.00;0.00'
COUNT_FORMAT = '#,##0'
LOT_FORMAT = '#,##0.0000'
CHECK_FORMAT = '0.00;[Red]-0.00;0.00'


def _text(value: object) -> str:
    return str(value or "").strip()


def _number(value: object) -> float:
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def _integer(value: object) -> int:
    try:
        return int(value or 0)
    except (TypeError, ValueError):
        return 0


def _joined(values: object, separator: str = "、") -> str:
    if not isinstance(values, (list, tuple, set)):
        return _text(values)
    return separator.join(_text(value) for value in values if _text(value))


def _currency(values: Iterable[object]) -> str:
    currencies = sorted({_text(value) for value in values if _text(value)})
    if len(currencies) == 1:
        return currencies[0]
    return "多币种" if currencies else ""


def _base_workbook(title: str) -> Workbook:
    workbook = Workbook()
    workbook.properties.title = title
    workbook.properties.creator = "K_desk"
    workbook.properties.subject = "账户自动化收益分析导出"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    return workbook


def _configure_sheet(sheet) -> None:
    sheet.sheet_view.showGridLines = False
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_view.zoomScale = 90


def _title(sheet, text: str, last_column: int) -> None:
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_column)
    cell = sheet.cell(1, 1, text)
    cell.fill = TITLE_FILL
    cell.font = Font(name="Microsoft YaHei", color="FFFFFF", bold=True, size=16)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    sheet.row_dimensions[1].height = 30


def _metadata(sheet, account: str, filters: dict, refreshed_at: str, exported_at: str, last_column: int) -> int:
    values = [
        ("查询账号", account),
        ("平台筛选", _text(filters.get("platform")) or "全部平台"),
        ("服务器筛选", _text(filters.get("server")) or "全部服务器"),
        ("数据刷新时间", refreshed_at or "-"),
        ("报表导出时间", exported_at or refreshed_at or "-"),
    ]
    row = 3
    for index, (label, value) in enumerate(values):
        column = 1 + index * 3
        if column + 1 > last_column:
            row += 1
            column = 1 + (index - 4) * 3
        sheet.cell(row, column, label)
        sheet.cell(row, column + 1, value)
        sheet.cell(row, column).fill = LABEL_FILL
        sheet.cell(row, column).font = HEADER_FONT
        sheet.cell(row, column + 1).font = BODY_FONT
        sheet.cell(row, column + 1).number_format = "@"
    return row + 2


def _section(sheet, row: int, text: str, last_column: int) -> int:
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_column)
    cell = sheet.cell(row, 1, text)
    cell.fill = SECTION_FILL
    cell.font = WHITE_FONT
    cell.alignment = Alignment(horizontal="left", vertical="center")
    sheet.row_dimensions[row].height = 22
    return row + 1


def _write_table(
    sheet,
    start_row: int,
    headers: list[str],
    rows: list[list[object]],
    *,
    table_name: str,
    account_columns: set[int] | None = None,
    count_columns: set[int] | None = None,
    lot_columns: set[int] | None = None,
    money_columns: set[int] | None = None,
    check_columns: set[int] | None = None,
    current_account_column: int | None = None,
) -> int:
    account_columns = account_columns or set()
    count_columns = count_columns or set()
    lot_columns = lot_columns or set()
    money_columns = money_columns or set()
    check_columns = check_columns or set()
    for column, header in enumerate(headers, 1):
        cell = sheet.cell(start_row, column, header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=THIN_GRAY)
    sheet.row_dimensions[start_row].height = 30
    for row_index, values in enumerate(rows, start_row + 1):
        for column, value in enumerate(values, 1):
            cell = sheet.cell(row_index, column, value)
            cell.font = BODY_FONT
            cell.alignment = Alignment(
                horizontal="right" if column in count_columns | lot_columns | money_columns | check_columns else "left",
                vertical="top",
                wrap_text=column not in account_columns | count_columns | lot_columns | money_columns | check_columns,
            )
            if column in account_columns:
                cell.value = _text(value)
                cell.number_format = "@"
            elif column in count_columns:
                cell.number_format = COUNT_FORMAT
            elif column in lot_columns:
                cell.number_format = LOT_FORMAT
            elif column in money_columns:
                cell.number_format = MONEY_FORMAT
            elif column in check_columns:
                cell.number_format = CHECK_FORMAT
        if current_account_column and _text(values[current_account_column - 1]) == "是":
            for column in range(1, len(headers) + 1):
                sheet.cell(row_index, column).fill = NOTE_FILL
    end_row = start_row + max(len(rows), 1)
    if rows:
        reference = f"A{start_row}:{get_column_letter(len(headers))}{end_row}"
        table = Table(displayName=table_name, ref=reference)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)
        sheet.auto_filter.ref = reference
    else:
        sheet.merge_cells(start_row=start_row + 1, start_column=1, end_row=start_row + 1, end_column=len(headers))
        empty_cell = sheet.cell(start_row + 1, 1, "当前查询范围没有可导出的数据")
        empty_cell.font = MUTED_FONT
        empty_cell.alignment = Alignment(horizontal="center", vertical="center")
        empty_cell.fill = PatternFill("solid", fgColor="F2F2F2")
    sheet.freeze_panes = f"A{start_row + 1}"
    if rows:
        for column in money_columns:
            letter = get_column_letter(column)
            target = f"{letter}{start_row + 1}:{letter}{end_row}"
            sheet.conditional_formatting.add(
                target,
                CellIsRule(operator="greaterThan", formula=["0"], fill=POSITIVE_FILL),
            )
            sheet.conditional_formatting.add(
                target,
                CellIsRule(operator="lessThan", formula=["0"], fill=NEGATIVE_FILL),
            )
        for column in check_columns:
            letter = get_column_letter(column)
            target = f"{letter}{start_row + 1}:{letter}{end_row}"
            sheet.conditional_formatting.add(
                target,
                CellIsRule(operator="notEqual", formula=["0"], fill=NEGATIVE_FILL),
            )
    return end_row


def _set_widths(sheet, widths: list[float]) -> None:
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(index)].width = min(max(width, 8), 42)


def _write_notes(
    sheet,
    title: str,
    definition: str,
    formula_note: str,
    errors: list[str],
    limitations: list[str],
) -> None:
    _configure_sheet(sheet)
    _title(sheet, title, 4)
    rows = [
        ("数据口径", definition or "-"),
        ("收益公式", formula_note),
        ("账户与币种", "账户号按文本保存；USC 金额按页面相同规则折算为 USD，并在明细中标识。"),
        ("数据权限", "报表仅使用只读查询结果，不修改 MT4、MT5、CRM 或本地台账。"),
        ("查询错误", "；".join(errors) if errors else "无"),
        ("数据限制", "；".join(dict.fromkeys(limitations)) if limitations else "无"),
    ]
    sheet.cell(3, 1, "项目")
    sheet.cell(3, 2, "说明")
    for cell in sheet[3][:2]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for row_index, (label, value) in enumerate(rows, 4):
        sheet.cell(row_index, 1, label).font = HEADER_FONT
        sheet.cell(row_index, 1).fill = LABEL_FILL
        sheet.cell(row_index, 2, value).font = BODY_FONT
        sheet.cell(row_index, 2).alignment = Alignment(wrap_text=True, vertical="top")
        sheet.row_dimensions[row_index].height = 34 if len(value) > 45 else 22
    sheet.freeze_panes = "A4"
    _set_widths(sheet, [18, 100, 12, 12])


def _workbook_bytes(workbook: Workbook) -> bytes:
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def _safe_master_sheet_name(account: str, index: int, used: set[str]) -> str:
    base = re.sub(r"[\\/*?:\[\]]", "_", f"单主_{account}")[:31] or f"单主_{index}"
    candidate = base
    suffix = 2
    while candidate.casefold() in used:
        tail = f"_{suffix}"
        candidate = f"{base[:31 - len(tail)]}{tail}"
        suffix += 1
    used.add(candidate.casefold())
    return candidate


def build_copy_profit_report(
    copy_payload: dict,
    signal_payload: dict | None,
    filters: dict,
    *,
    exported_at: str = "",
) -> bytes:
    origins = copy_payload.get("origins") if isinstance(copy_payload.get("origins"), list) else []
    workbook = _base_workbook("单主跟单收益报表")
    summary_sheet = workbook.active
    summary_sheet.title = "单主汇总"
    _configure_sheet(summary_sheet)
    summary_headers = ["单主账号", "平台", "服务器", "跟单账户", "跟单订单", "跟单手数", "跟单总盈亏", "币种"]
    _title(summary_sheet, "单主跟单总盈亏汇总", len(summary_headers))
    summary_rows = []
    for origin in origins:
        totals = origin.get("followerSummary") or {}
        summary_rows.append([
            _text(origin.get("account")), _text(origin.get("platform")), _text(origin.get("server")),
            _integer(totals.get("accounts")), _integer(totals.get("orders")), _number(totals.get("volume")),
            _number(totals.get("netProfit")), _text(totals.get("currency")),
        ])
    _write_table(
        summary_sheet,
        3,
        summary_headers,
        summary_rows,
        table_name="CopyMasterSummary",
        account_columns={1},
        count_columns={4, 5},
        lot_columns={6},
        money_columns={7},
    )
    _set_widths(summary_sheet, [16, 12, 22, 13, 13, 13, 18, 12])

    used_sheet_names = {"单主汇总".casefold()}
    for index, origin in enumerate(origins, 1):
        master_account = _text(origin.get("account"))
        totals = origin.get("followerSummary") or {}
        sheet = workbook.create_sheet(_safe_master_sheet_name(master_account, index, used_sheet_names))
        _configure_sheet(sheet)
        _title(sheet, f"单主 {master_account} 跟单收益", 15)

        metrics = [
            ("平台 / 服务器", " / ".join(filter(None, [_text(origin.get("platform")), _text(origin.get("server"))])) or "-"),
            ("跟单账户", _integer(totals.get("accounts"))),
            ("跟单订单", _integer(totals.get("orders"))),
            ("跟单手数", _number(totals.get("volume"))),
            ("跟单总盈亏", _number(totals.get("netProfit"))),
            ("币种", _text(totals.get("currency")) or "-"),
        ]
        for metric_index, (label, value) in enumerate(metrics):
            column = 1 + metric_index * 2
            sheet.cell(3, column, label).fill = LABEL_FILL
            sheet.cell(3, column, label).font = HEADER_FONT
            sheet.cell(4, column, value).font = Font(name="Microsoft YaHei", bold=True, size=12, color="17365D")
            if label == "跟单账户" or label == "跟单订单":
                sheet.cell(4, column).number_format = COUNT_FORMAT
            elif label == "跟单手数":
                sheet.cell(4, column).number_format = LOT_FORMAT
            elif label == "跟单总盈亏":
                sheet.cell(4, column).number_format = MONEY_FORMAT

        row = _section(sheet, 6, "跟单收益汇总", 15)
        follower_headers = [
            "跟单账号", "平台", "服务器", "匹配源单", "跟单订单", "手数", "毛盈亏", "手续费/Fee",
            "利息/Swap", "税费", "净盈亏", "币种", "品种", "首次交易", "最后交易",
        ]
        follower_rows = [
            [
                _text(member.get("account")), _text(member.get("platform") or origin.get("platform")),
                _text(member.get("server") or origin.get("server")), _integer(member.get("matchedSourceOrders")),
                _integer(member.get("orders")), _number(member.get("volume")), _number(member.get("grossProfit")),
                _number(member.get("commission")), _number(member.get("swap")), _number(member.get("taxes")),
                _number(member.get("netProfit")), _text(member.get("displayCurrency") or member.get("currency")),
                _joined(member.get("symbols")), _text(member.get("firstTime")), _text(member.get("lastTime")),
            ]
            for member in (origin.get("followers") or [])
        ]
        follower_end = _write_table(
            sheet,
            row,
            follower_headers,
            follower_rows,
            table_name=f"CopyMasterFollowers{index}",
            account_columns={1},
            count_columns={4, 5},
            lot_columns={6},
            money_columns={7, 8, 9, 10, 11},
        )

        order_section = _section(sheet, follower_end + 2, "跟单订单明细", 15)
        order_headers = [
            "跟单账号", "平台", "服务器", "订单 / Position", "源订单号", "品种", "开仓时间", "平仓时间",
            "手数", "毛盈亏", "手续费/Fee", "利息/Swap", "税费", "净盈亏", "币种",
        ]
        follower_orders = origin.get("followerOrders") if isinstance(origin.get("followerOrders"), list) else []
        follower_orders = sorted(
            follower_orders,
            key=lambda item: (_text(item.get("account")), _text(item.get("openTime")), _text(item.get("ticket"))),
        )
        order_rows = [
            [
                _text(order.get("account")), _text(order.get("platform") or origin.get("platform")),
                _text(order.get("server") or origin.get("server")), _text(order.get("ticket") or _joined(order.get("tickets"))),
                _joined(order.get("matchedSourceOrderIds")), _text(order.get("symbol") or _joined(order.get("symbols"))),
                _text(order.get("openTime")), _text(order.get("closeTime")), _number(order.get("volume")),
                _number(order.get("grossProfit")), _number(order.get("commission")), _number(order.get("swap")),
                _number(order.get("taxes")), _number(order.get("netProfit")),
                _text(order.get("displayCurrency") or order.get("currency")),
            ]
            for order in follower_orders
        ]
        _write_table(
            sheet,
            order_section,
            order_headers,
            order_rows,
            table_name=f"CopyMasterOrders{index}",
            account_columns={1, 4, 5},
            lot_columns={9},
            money_columns={10, 11, 12, 13, 14},
        )
        _set_widths(sheet, [15, 11, 22, 20, 22, 16, 20, 20, 12, 15, 15, 14, 12, 15, 11])

    workbook.active = 0
    return _workbook_bytes(workbook)


def build_ea_profit_report(payload: dict, filters: dict, *, exported_at: str = "") -> bytes:
    account = _text(payload.get("account"))
    workbook = _base_workbook(f"EA 收益报表 - {account}")
    summary_sheet = workbook.active
    summary_sheet.title = "EA汇总"
    summary_headers = [
        "EA Comment", "EA标识", "数据库", "平台", "服务器", "账户数", "订单数", "手数",
        "毛盈亏", "手续费/Fee", "利息/Swap", "税费", "净盈亏", "币种", "当前账号订单",
        "当前账号净盈亏", "首次交易", "最后交易", "匹配规则",
    ]
    groups = payload.get("groups") if isinstance(payload.get("groups"), list) else []
    summary_sheet.append(summary_headers)
    for group in groups:
        totals = group.get("totals") or {}
        summary_sheet.append([
            ("[可能是跟单路由] " if group.get("classification") == "possible_copy_route" else "")
            + _text(group.get("comment")), _text(group.get("expertId")),
            _text(group.get("database") or _joined(group.get("databases"), " / ")),
            _text(group.get("platform")), _text(group.get("server")),
            _integer(totals.get("accounts")), _integer(totals.get("orders")), _number(totals.get("volume")),
            _number(totals.get("grossProfit")),
            _number(totals.get("commission")), _number(totals.get("swap")), _number(totals.get("taxes")),
            _number(totals.get("netProfit")), _text(totals.get("currency")), _integer(group.get("currentOrders")),
            _number(group.get("currentNetProfit")), _text(group.get("firstTime")), _text(group.get("lastTime")),
            _text(group.get("matchRule")),
        ])

    detail_sheet = workbook.create_sheet("EA明细")
    detail_headers = [
        "EA", "EA标识", "匹配线索", "数据库", "平台", "服务器", "账户", "当前账号", "平仓订单", "手数", "毛盈亏", "手续费/Fee",
        "利息/Swap", "税费", "净盈亏", "币种", "USC折算", "品种", "首次交易", "最后交易", "样例订单号",
    ]
    detail_sheet.append(detail_headers)
    for group in groups:
        for member in (group.get("members") or []):
            detail_sheet.append([
            ("[可能是跟单路由] " if group.get("classification") == "possible_copy_route" else "")
            + _text(group.get("comment")), _joined(member.get("expertIds")),
            _joined(member.get("matchClues"), "；") or _text(member.get("matchClue")),
            _text(member.get("database") or group.get("database")),
            _text(member.get("platform") or group.get("platform")), _text(member.get("server") or group.get("server")),
            _text(member.get("account")), "是" if member.get("isCurrentAccount") else "否",
            _integer(member.get("orders")), _number(member.get("volume")), _number(member.get("grossProfit")),
            _number(member.get("commission")), _number(member.get("swap")), _number(member.get("taxes")),
            _number(member.get("netProfit")), _text(member.get("currency")),
            "是" if member.get("isCentAccount") else "否", _joined(member.get("symbols")),
            _text(member.get("firstTime")), _text(member.get("lastTime")), _joined(member.get("tickets")),
            ])
    workbook.active = 0
    return _workbook_bytes(workbook)
