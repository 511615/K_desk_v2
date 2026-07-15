from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from kdesk.domain.ledger import LEGACY_HEADERS, LEGACY_HISTORY_HEADERS, LedgerRecord, clean
from kdesk.infrastructure.database import Database


def _sheet_rows(path: Path, sheet_name: str) -> list[dict[str, str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        workbook.close()
        return []
    sheet = workbook[sheet_name]
    values = list(sheet.iter_rows(values_only=True))
    workbook.close()
    if not values:
        return []
    headers = [clean(value) for value in values[0]]
    return [
        {header: clean(value) for header, value in zip(headers, row, strict=False) if header}
        for row in values[1:]
        if any(clean(value) for value in row)
    ]


def preview_import(path: Path, database: Database) -> dict:
    records = _sheet_rows(path, "问题账户")
    existing = {item.record_id: item.to_legacy() for item in database.list_accounts()}
    added = 0
    changed = 0
    unchanged = 0
    conflicts: list[dict[str, str]] = []
    for row in records:
        record = LedgerRecord.from_mapping(row)
        current = existing.get(record.record_id)
        if current is None:
            added += 1
        elif any(clean(current.get(key)) != clean(record.to_legacy().get(key)) for key in LEGACY_HEADERS):
            changed += 1
            if current.get("账号") != record.account:
                conflicts.append({"recordId": record.record_id, "databaseAccount": current.get("账号", ""), "excelAccount": record.account})
        else:
            unchanged += 1
    return {"records": len(records), "added": added, "changed": changed, "unchanged": unchanged, "conflicts": conflicts}


def import_workbook(path: Path, database: Database, *, allow_conflicts: bool = False) -> dict:
    preview = preview_import(path, database)
    if preview["conflicts"] and not allow_conflicts:
        raise ValueError("Excel 导入存在记录ID冲突，请先处理冲突")
    for row in _sheet_rows(path, "问题账户"):
        database.import_account(LedgerRecord.from_mapping(row))
    history_rows = _sheet_rows(path, "修改历史")
    for row in history_rows:
        database.import_history(row)
    return {**preview, "history": len(history_rows), "imported": True}


def export_workbook(path: Path, database: Database) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    accounts_sheet = workbook.active
    accounts_sheet.title = "问题账户"
    accounts_sheet.append(LEGACY_HEADERS)
    for record in database.list_accounts():
        payload = record.to_legacy()
        accounts_sheet.append([payload.get(header, "") for header in LEGACY_HEADERS])

    history_sheet = workbook.create_sheet("修改历史")
    history_sheet.append(LEGACY_HISTORY_HEADERS)
    for record in database.list_accounts():
        for row in reversed(database.history(record.record_id)):
            history_sheet.append([row.get(header, "") for header in LEGACY_HISTORY_HEADERS])

    help_sheet = workbook.create_sheet("字段说明")
    help_sheet.append(["字段", "说明"])
    for header in LEGACY_HEADERS:
        help_sheet.append([header, "由 K_desk v2 SQLite 权威数据导出"])

    for sheet in (accounts_sheet, history_sheet, help_sheet):
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions

    workbook.save(path)
    return path


def export_audit_json(path: Path, result: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
