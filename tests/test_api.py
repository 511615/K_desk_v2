from __future__ import annotations

import gc
import time
from dataclasses import replace
from io import BytesIO
from pathlib import Path

import kuzu
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from kdesk.api.account_app import (
    RELATIONSHIP_DISCOVERY_TIMEOUT_SECONDS,
    RELATIONSHIP_SOURCE_TIMEOUT_SECONDS,
    create_account_app,
)
from kdesk.api.kline_app import create_kline_app
from kdesk.application.relationship_process import IsolatedRelationshipRiskBuilder
from kdesk.infrastructure.legacy_bridge import LegacyBridge
from kdesk.settings import Settings


def make_test_settings(tmp_path: Path) -> Settings:
    runtime = tmp_path / "runtime"
    return Settings(
        root=tmp_path,
        profile="test",
        host="127.0.0.1",
        account_port=8877,
        kline_port=8866,
        runtime_dir=runtime,
        database_path=runtime / "kdesk.sqlite",
        queue_database_path=runtime / "jobs.sqlite",
        artifact_dir=runtime / "artifacts",
        upload_dir=runtime / "uploads",
        log_dir=runtime / "logs",
        legacy_root=tmp_path / "legacy",
        legacy_output=tmp_path / "legacy_output",
        legacy_trade_database=tmp_path / "trades.sqlite",
        bootstrap_xlsx=runtime / "import" / "problematic_accounts.xlsx",
        legacy_compat_dir=runtime / "legacy_compat",
        frontend_dist=tmp_path / "frontend" / "dist",
        ui_mode="vue",
    )


def test_account_relationship_runtime_allows_long_background_reads_without_a_global_deadline(tmp_path: Path) -> None:
    app = create_account_app(make_test_settings(tmp_path))

    assert RELATIONSHIP_DISCOVERY_TIMEOUT_SECONDS is None
    assert RELATIONSHIP_SOURCE_TIMEOUT_SECONDS == 120.0
    assert app.state.relationship_risk._discovery_timeout_seconds is None
    assert app.state.relationship_network._source_timeout_seconds == 120.0


def test_production_relationship_runtime_uses_a_disposable_process(tmp_path: Path) -> None:
    settings = replace(make_test_settings(tmp_path), profile="prod")

    app = create_account_app(settings)

    assert isinstance(app.state.relationship_risk, IsolatedRelationshipRiskBuilder)
    assert app.state.relationship_expansion._risk_builder is app.state.relationship_risk
    assert app.state.relationship_risk._process_timeout_seconds == 45.0


def test_health_and_ledger_api_are_isolated(tmp_path: Path) -> None:
    app = create_account_app(make_test_settings(tmp_path))
    with TestClient(app) as client:
        health = client.get("/health/ready").json()
        assert health["ok"] is True
        assert health["relationshipExpansion"] == {
            "residentJobs": 0,
            "runningJobs": 0,
            "queuedJobs": 0,
            "completedJobs": 0,
            "maxResidentJobs": 3,
            "cacheSeconds": 90.0,
        }
        saved = client.post("/api/accounts/mark", json={"account": "302360", "action": "M", "status": "观察中"})
        assert saved.status_code == 200
        ledger = client.get("/api/accounts/by-login/302360/ledger").json()
        assert ledger["marked"] is True
        assert ledger["record"]["建议动作"] == "M"


def test_kline_upload_page_chains_inspection_to_generation(tmp_path: Path) -> None:
    app = create_kline_app(make_test_settings(tmp_path))
    with TestClient(app) as client:
        response = client.get("/")

    assert response.status_code == 200
    assert "startGeneration" in response.text
    assert "/api/jobs/'+inspect.id+'/generate" in response.text
    assert "打开生成图表" in response.text


def test_account_kline_job_accepts_bounded_recent_order_window(tmp_path: Path) -> None:
    app = create_account_app(make_test_settings(tmp_path))
    with TestClient(app) as client:
        response = client.post(
            "/api/kline/generate-from-db",
            json={
                "account": "302360",
                "platform": "MT5",
                "server": "DBG MT5",
                "recentOrders": 300,
                "cacheVersion": "2026-08-24 10:00:00",
            },
        )
        job = client.get(f"/api/kline/jobs/{response.json()['job']['id']}").json()["job"]

    assert response.status_code == 200
    assert job["payload"]["recentOrders"] == 300
    assert job["payload"]["cacheVersion"] == "2026-08-24 10:00:00"


def test_api_meta_exposes_governed_build_information(tmp_path: Path) -> None:
    app = create_account_app(make_test_settings(tmp_path))
    with TestClient(app) as client:
        payload = client.get("/api/meta").json()
    assert payload["version"] == "2.1.4"
    assert payload["gitSha"]
    assert payload["buildTime"]
    assert payload["schemaRevision"] in {"unversioned", "uninitialized", "0001"}
    assert payload["featureRegistryVersion"]
    assert payload["featureCount"] >= 10
    assert payload["compatibilityLevel"] == "legacy-account-v1"
    assert payload["sourceRoot"]
    assert payload["pythonExecutable"]
    assert payload["branch"] in {"main", "dev", "detached"} or payload["branch"].startswith("feature/")
    assert payload["defaultRoutes"] == {
        "kuzuRisk": "focus-force",
        "kuzuGalaxy": "graph_type=galaxy",
    }


def test_account_validation_rejects_unsafe_path_data(tmp_path: Path) -> None:
    app = create_account_app(make_test_settings(tmp_path))
    with TestClient(app) as client:
        response = client.post("/api/accounts/mark", json={"account": "../bad"})
        assert response.status_code == 400


def test_copy_time_range_validation_returns_bad_request(tmp_path: Path, monkeypatch) -> None:
    def fake_call(_self, name, *args):
        assert name == "account_copy_origins_payload"
        raise ValueError("跟单开始时间不能晚于结束时间")

    monkeypatch.setattr(LegacyBridge, "call", fake_call)
    app = create_account_app(make_test_settings(tmp_path))
    with TestClient(app) as client:
        response = client.get(
            "/api/accounts/by-login/302360/copy-origins"
            "?start=2026-08-02%2000:00:00&end=2026-08-01%2000:00:00"
        )
    assert response.status_code == 400
    assert response.json()["error"] == "跟单开始时间不能晚于结束时间"


def test_account_detail_always_uses_legacy_page(tmp_path: Path, monkeypatch) -> None:
    monkeypatch.setattr(LegacyBridge, "account_page", lambda _self, login: f"<html><body>legacy-account:{login}</body></html>")
    app = create_account_app(make_test_settings(tmp_path))
    with TestClient(app) as client:
        response = client.get("/account/302360?platform=MT5&server=DBG%20MT5")
        assert response.status_code == 200
        assert "legacy-account:302360" in response.text


def test_account_inline_kline_is_served_by_account_service_not_the_job_api(tmp_path: Path, monkeypatch) -> None:
    def fake_call(_self, name, *args):
        assert name == "account_inline_kline_html"
        assert args == ("302360", {"platform": "MT5", "server": "DBG MT5", "recentOrders": 300})
        return "<html><body>direct-lightweight-kline</body></html>"

    monkeypatch.setattr(LegacyBridge, "call", fake_call)
    app = create_account_app(make_test_settings(tmp_path))
    with TestClient(app) as client:
        response = client.get(
            "/api/accounts/by-login/302360/inline-kline?platform=MT5&server=DBG%20MT5&recentOrders=300"
        )

    assert response.status_code == 200
    assert response.headers["cache-control"] == "private, max-age=60"
    assert "direct-lightweight-kline" in response.text


def test_historical_funds_api_replays_read_only_source_facts(tmp_path: Path, monkeypatch) -> None:
    def fake_call(_self, name, *args):
        assert name == "account_historical_funds_source_payload"
        return {
            "available": True,
            "account": "302360",
            "platform": "MT4",
            "server": "DBG MT4 CN1",
            "source": "fixture",
            "currency": "USD",
            "moneyScale": 1,
            "coverage": {"eventRows": 2, "dailyAnchors": 1, "completeHistory": True},
            "anchors": [{"timestamp": "2026-01-01 00:00:00", "balance": 100, "credit": 20, "equity": 120}],
            "events": [
                {"id": "1", "timestamp": "2026-01-01 01:00:00", "CMD": 6, "PROFIT": 100, "COMMENT": "DEP-EO-1"},
                {"id": "2", "timestamp": "2026-01-01 02:00:00", "CMD": 7, "PROFIT": 10, "COMMENT": "BONUS"},
            ],
        }

    monkeypatch.setattr(LegacyBridge, "call", fake_call)
    app = create_account_app(make_test_settings(tmp_path))
    with TestClient(app) as client:
        response = client.get("/api/accounts/by-login/302360/historical-funds?platform=MT4&server=DBG%20MT4%20CN1")

    assert response.status_code == 200
    payload = response.json()
    assert payload["ok"] is True
    assert payload["available"] is True
    assert payload["summary"]["externalDeposit"] == 100
    assert payload["summary"]["bonusGranted"] == 10
    assert payload["events"][-1]["balance"] == 200
    assert payload["events"][-1]["credit"] == 30


def test_historical_funds_api_returns_a_readable_failure_without_database_detail(tmp_path: Path, monkeypatch) -> None:
    def failing_call(_self, _name, *_args):
        raise RuntimeError("source timeout should stay in server logs")

    monkeypatch.setattr(LegacyBridge, "call", failing_call)
    app = create_account_app(make_test_settings(tmp_path))
    with TestClient(app) as client:
        response = client.get("/api/accounts/by-login/302360/historical-funds?platform=MT5&server=DBG%20MT5")

    assert response.status_code == 503
    payload = response.json()
    assert payload["ok"] is False
    assert payload["available"] is False
    assert payload["error"] == "历史资金数据查询失败，请检查数据库连接"
    assert "timeout" not in payload["error"]


def test_relationship_network_returns_kuzu_scored_evidence_with_partial_source_coverage(tmp_path: Path, monkeypatch) -> None:
    calls: list[tuple[str, tuple]] = []

    def fake_call(_self, name, *args):
        calls.append((name, args))
        if name == "account_relationship_core_payload":
            return {
                "riskPanels": {
                    "available": True,
                    "databaseStatus": "TA",
                    "sameName": [{
                        "account": "302361", "platform": "MT5", "server": "DBG MT5", "databaseStatus": "P",
                    }],
                }
            }
        if name == "account_login_ips_payload":
            return {"records": [{
                "ip": "203.0.113.42", "platform": "MT5", "server": "DBG MT5",
                "lastAccessAt": "2026-07-29 10:00:00", "firstSeenAt": "2026-07-01 10:00:00",
                "lastSeenAt": "2026-07-29 10:00:00", "geo": {"country": "CN"},
            }]}
        if name == "account_shared_last_ip_payload":
            return {"peers": [{"account": "302365", "platform": "MT5", "server": "DBG MT5", "ip": "203.0.113.42", "lastAccessAt": "2026-07-29 10:00:00"}], "coverage": [{"source": "sharedLastIp", "status": "available", "reason": ""}]}
        if name == "account_shared_cid_payload":
            return {"peers": [], "coverage": [{"source": "sharedCid", "status": "available", "reason": "当前账户 CID 为空或 0"}]}
        if name == "account_ea_comment_profit_payload":
            return {"groups": [{
                "comment": "GoldBot", "platform": "MT5", "server": "DBG MT5",
                "classificationLabel": "EA", "matchRule": "Comment + ExpertID",
                "members": [{
                    "account": "302362", "platform": "MT5", "server": "DBG MT5",
                    "orders": 8, "netProfit": 42.5, "currency": "USD", "matchClue": "Comment matched",
                }],
            }]}
        if name == "account_copy_origins_payload":
            return {"origins": [{
                "account": "302363", "platform": "MT5", "server": "DBG MT5", "matchedOrders": 4,
                "orders": 4, "netProfit": 11.2, "currency": "USD", "followers": [{
                    "account": "302364", "platform": "MT5", "server": "DBG MT5", "orders": 4,
                    "netProfit": 9.8, "currency": "USD",
                }],
            }]}
        if name == "account_copy_group_profit_payload":
            raise RuntimeError("copy group source temporarily unavailable")
        if name == "account_crm_ib_relationship_payload":
            return {"records": []}
        raise AssertionError(name)

    monkeypatch.setattr(LegacyBridge, "call", fake_call)
    app = create_account_app(make_test_settings(tmp_path))
    with TestClient(app) as client:
        response = client.get("/api/accounts/by-login/302360/relationship-network?platform=MT5&server=DBG%20MT5")
        payload = response.json()
        deadline = time.monotonic() + 1
        while payload.get("inProgress") and time.monotonic() < deadline:
            time.sleep(0.01)
            payload = client.get("/api/accounts/by-login/302360/relationship-network?platform=MT5&server=DBG%20MT5").json()

    assert response.status_code == 200
    assert payload["inProgress"] is False
    assert payload["account"] == "302360"
    assert payload["source"] == "kuzu-request-projection"
    assert payload["threshold"] == 12
    assert {item["id"] for item in payload["relationTypes"]} == {
        "same_crm_user", "login_ip", "client_id", "ea_feature", "copy_order", "copy_group", "rebate",
        "crm_owner", "direct_ib", "ib_owned_account", "ib_direct_account", "ib_identity", "ib_direct_rebate", "top_ib_group",
        "toxic_sync_same", "toxic_sync_opposite",
    }
    assert {item["type"] for item in payload["relationships"]} == {
        "same_crm_user", "login_ip", "ea_feature", "copy_order",
    }
    assert payload["summary"]["entityCount"] >= 7
    assert any(item["source"] == "copyGroups" and item["status"] == "failed" for item in payload["coverage"])
    assert "调查优先级" in payload["limitations"][0]
    assert all("score" in item and "riskColor" in item for item in payload["entities"])
    assert next(item for item in payload["entities"] if item["label"] == "302360")["databaseStatus"] == "TA"
    assert next(item for item in payload["entities"] if item["label"] == "302361")["databaseStatus"] == "P"
    presentation = payload["presentationGraph"]
    assert presentation["modelVersion"] == "relationship-entity-v1"
    assert presentation["subjectId"] == "account:302360|MT5|DBG MT5"
    assert next(item for item in presentation["entities"] if item["label"] == "302360")["databaseStatus"] == "TA"
    assert next(item for item in presentation["entities"] if item["label"] == "302361")["databaseStatus"] == "P"
    assert all("localAction" not in item for item in payload["entities"])
    assert {name for name, _args in calls} == {
        "account_relationship_core_payload", "account_copy_origins_payload",
        "account_copy_group_profit_payload", "account_ea_comment_profit_payload", "account_crm_ib_relationship_payload",
        "account_shared_last_ip_payload", "account_shared_cid_payload",
    }
    relationship_sources = {
        "account_relationship_core_payload", "account_copy_origins_payload",
        "account_copy_group_profit_payload", "account_ea_comment_profit_payload", "account_crm_ib_relationship_payload",
    }
    for name, args in calls:
        if name in relationship_sources:
            assert args[1]["_relationship"] == "1"


def test_kuzu_risk_legacy_galaxy_requires_explicit_graph_type(tmp_path: Path) -> None:
    app = create_account_app(make_test_settings(tmp_path))

    with TestClient(app) as client:
        page = client.get("/kuzu-risk?account=302360&platform=MT5&server=DBG%20MT5&graph_type=galaxy")

    assert page.status_code == 200
    assert "/api/accounts/by-login/" in page.text
    assert "关系路径说明" in page.text
    assert "问题账户的直属上级 IB 本人名下交易账户" in page.text
    assert "同名账户" in page.text
    assert "当前 CID 相同" in page.text
    assert "原始依据：" not in page.text
    assert "directSubjectEdge" in page.text
    assert "ib_direct_account','ib_direct_rebate" in page.text
    assert "ib_direct_rebate" in page.text
    assert "graphNodes" in page.text
    assert "function queuePoll" in page.text
    assert "queuePoll=function(delay=2000)" in page.text
    assert "document.hidden" in page.text
    assert "pagehide" in page.text
    assert "data.inProgress" in page.text
    assert "后台扩散中：已处理" in page.text
    assert "关系扩散扫描中" in page.text
    assert "MutationObserver" in page.text
    assert "repeatCount=\"indefinite\"" in page.text
    assert "pointerEvents:'none'" in page.text
    assert "function positionRadar" in page.text
    assert "radarSweep.setAttribute('transform'" in page.text
    assert "radarSvg.setAttribute('viewBox'" in page.text
    assert "radarFan.setAttribute('d'" in page.text
    assert "const NODE_SCALE=2" in page.text
    assert "hitRadius=26/Math.max(view.scale,.1)" in page.text
    assert 'id="riskTable"' in page.text
    assert "function riskStatusAccounts" in page.text
    assert "function renderRiskTable" in page.text
    assert "['T','TA','A'].includes(databaseStatus(node))" in page.text
    assert "跟单订单匹配（开仓/平仓）" in page.text
    assert "Toxic 同向开平仓时间匹配" in page.text
    assert "Toxic 反向开平仓时间匹配" in page.text
    assert "包含 Toxic 同向/反向开平仓时间匹配（较慢）" in page.text
    assert "copy_order:'同步订单'" not in page.text
    assert "包含 Toxic 同步订单（较慢）" not in page.text
    assert page.text.count("const relationNames=") == 0
    assert "copyInspector" in page.text
    assert "fetchCopyInspection" in page.text
    assert "跟单订单明细" in page.text
    assert "该单主的全部跟单账户" in page.text
    assert "relationship-network" in page.text
    assert 'id="includeToxic"' in page.text
    assert "rangeStart.id='rangeStart'" in page.text
    assert "rangeEnd.id='rangeEnd'" in page.text
    assert '留空=全历史' in page.text
    assert "query.set('start',rangeStart.value.replace('T',' '))" in page.text
    assert "query.set('end',rangeEnd.value.replace('T',' '))" in page.text
    assert "if(includeToxic.checked)query.set('include_toxic','true')" in page.text
    assert 'id="overview"' in page.text
    assert "function accountDepth" in page.text
    assert "function relationEdgeLabel" in page.text
    assert "仅显示当前账户与主账户的路径和直接关系" in page.text
    assert "function localAction" in page.text
    assert "function drawActionBadge" in page.text
    assert "直属上级 IB 本人账户" in page.text
    assert "function renderOverview" in page.text
    assert "function ringLayout" in page.text
    assert "全部发现节点均单独布局" in page.text
    assert "function relationshipGroup" in page.text
    assert "function drawGroupBand" in page.text
    assert "function drawGroupLabel" in page.text
    assert "function selectedBranch" in page.text
    assert "function relationshipCluster" in page.text
    assert "function isDirectedRelation" in page.text
    assert "function drawRelationEdge" in page.text
    assert "relationshipGroup(to)" in page.text
    assert "selectedEdgeKey" in page.text
    assert "expandedRelationGroups" in page.text
    assert "展开当前关系群组" in page.text
    assert "合并当前关系群组" in page.text
    assert "群组成员" in page.text
    assert "relationshipGroup(target)" in page.text
    assert "function edgeCommunityKey" in page.text
    assert "relationHitEdges" in page.text
    assert "distanceToRelationEdge" in page.text
    assert "stopImmediatePropagation()" in page.text
    assert "function groupAnchorPoint" in page.text
    assert "groupMembers" in page.text
    assert "群落" in page.text
    assert "function galaxyRelationshipCommunities" in page.text
    assert "function galaxySameCrmCommunity" in page.text
    assert "function galaxyComponentVisualDepth" in page.text
    assert "function galaxyCommunityMemberships" in page.text
    assert "function galaxyOrbitOverlapBands" in page.text
    assert "orbitOnly:true" in page.text
    assert "span=.12" in page.text
    assert "nodes:[node]" in page.text
    assert "group.nodes.some(node=>galaxyCommunityMembershipCount(node.id)>1)" in page.text
    assert "if(group.componentMemberCount&&group.type==='same_crm_user')continue" in page.text
    assert "group.componentMemberCount+'账户'" in page.text
    assert "function drawIntersectingCommunities" not in page.text
    assert "function drawOrbitOverlapBands" not in page.text
    assert "band.lane" not in page.text
    assert "function relationRoute" in page.text
    assert "quadraticCurveTo" in page.text
    assert "routeLaneCursor" in page.text
    assert page.text.count("function renderOverview") == 1
    assert "function relationTheme" in page.text
    assert "function nodeShape" in page.text
    assert "最低缩放 10%" in page.text
    assert "function terminalState" in page.text
    assert "function drawTerminalBadge" in page.text
    assert "已核查叶节点" in page.text
    assert "白色虚线：当前选中（不改色）" in page.text
    assert "function renderDetail" in page.text
    assert "function checkedLeaf" in page.text
    assert "已核查，无新增账户" in page.text
    assert "当前账户细查" in page.text


def test_kuzu_risk_galaxy_uses_one_immutable_click_dispatcher(tmp_path: Path) -> None:
    app = create_account_app(make_test_settings(tmp_path))

    with TestClient(app) as client:
        page = client.get("/kuzu-risk?account=302360&platform=MT5&server=DBG%20MT5&graph_type=galaxy")

    assert page.status_code == 200
    assert "function galaxyRebuildHitFrame" in page.text
    assert "function galaxyPickHit" in page.text
    assert "function galaxyDispatchClick" in page.text
    assert "function galaxyVisualEndpointKey" in page.text
    assert "galaxyCanvas.addEventListener('click',galaxyDispatchClick,true)" in page.text
    assert "event.stopImmediatePropagation()" in page.text

    picker = page.text.split("function galaxyPickHit", 1)[1].split("function galaxyDispatchClick", 1)[0]
    dispatcher = page.text.split("function galaxyDispatchClick", 1)[1].split("const ungroupedSelectedBranch", 1)[0]
    assert "ringLayout(" not in picker
    assert "ringLayout(" not in dispatcher
    assert picker.index("frame.markers") < picker.index("frame.nodes")
    assert picker.index("frame.nodes") < picker.index("frame.groups")
    assert picker.index("frame.groups") < picker.index("frame.edges")
    assert "kind==='marker'" in dispatcher
    assert "kind==='node'" in dispatcher
    assert "kind==='group'" in dispatcher
    assert "kind==='edge'" in dispatcher
    assert "galaxyVisualEndpointKey(edge?.from)" in page.text
    assert "galaxyVisualEndpointKey(edge?.to)" in page.text


def test_kuzu_risk_galaxy_has_lazy_profile_and_relation_evidence_ui(tmp_path: Path) -> None:
    app = create_account_app(make_test_settings(tmp_path))
    with TestClient(app) as client:
        page = client.get("/kuzu-risk?account=302360&platform=MT5&server=DBG%20MT5&graph_type=galaxy")
    assert page.status_code == 200
    for marker in (
        "relationship-network/node-profile", "relationship-network/relation-detail",
        "AbortController", "inspectionRequestSequence", "查看账户详情",
        "高度关联账户", "关系证据", "时间同步线索，不等同于跟单",
        "inspectionPanel.hidden=true;document.querySelector('.detail')?.prepend(inspectionPanel)",
        "inspection-hero", "传播分", "已完成扩散 · 无新增账户",
    ):
        assert marker in page.text

    terminal_state = page.text.rsplit("function terminalState", 1)[1].split("// ACC-REL-001", 1)[0]
    assert "node.expansionState!=='expanded'" in terminal_state
    assert "!node.expansionEvidenceAvailable" in terminal_state


def test_relationship_inspection_endpoints_read_the_current_snapshot(tmp_path: Path, monkeypatch) -> None:
    app = create_account_app(make_test_settings(tmp_path))
    snapshot = {
        "revision": 7,
        "subjectId": "account:100",
        "entities": [
            {"id": "account:100", "type": "account", "label": "100", "isSubject": True, "databaseStatus": "M"},
            {
                "id": "account:101", "type": "account", "label": "101", "databaseStatus": "P", "hops": 1,
                "expandable": True, "expansionState": "expanded", "expansionEvidenceAvailable": True,
            },
        ],
        "relationships": [
            {"id": "same-crm:100:101", "source": "account:100", "target": "account:101", "type": "same_crm_user"},
        ],
        "coverage": [],
        "limitations": [],
        "inProgress": False,
    }
    monkeypatch.setattr(app.state.relationship_expansion, "get_or_start", lambda *_args: snapshot)
    monkeypatch.setattr(app.state.legacy, "call", lambda *_args: {})

    with TestClient(app) as client:
        profile = client.get(
            "/api/accounts/by-login/100/relationship-network/node-profile",
            params={"node_id": "account:101", "platform": "MT5", "server": "AC CN MT5"},
        )
        relation = client.get(
            "/api/accounts/by-login/100/relationship-network/relation-detail",
            params={
                "edge_id": "same-crm:100:101",
                "platform": "MT5",
                "server": "AC CN MT5",
                "start": "2026-05-24 00:00:00",
                "end": "2026-08-24 23:59:59",
                "job_id": "investigation-7",
            },
        )
        stale = client.get(
            "/api/accounts/by-login/100/relationship-network/relation-detail",
            params={"edge_id": "same-crm:100:101", "snapshot_version": 6},
        )

    assert profile.status_code == 200
    assert profile.json()["account"]["database_status"] == "P"
    assert profile.json()["account"]["expansion_state"] == "expanded"
    assert profile.json()["account"]["expansion_evidence_available"] is True
    assert profile.json()["coverage"]["start"]
    assert relation.status_code == 200
    assert relation.json()["relations"][0]["business_name"] == "同名账户"
    assert relation.json()["relations"][0]["time_range"] == {
        "start": "2026-05-24 00:00:00",
        "end": "2026-08-24 23:59:59",
    }
    assert relation.json()["job_id"] == "investigation-7"
    assert "user_id" not in relation.text
    assert stale.status_code == 409


def test_kuzu_risk_galaxy_locator_is_independent_risk_colored_and_clickable(tmp_path: Path) -> None:
    app = create_account_app(make_test_settings(tmp_path))

    with TestClient(app) as client:
        page = client.get("/kuzu-risk?account=302360&platform=MT5&server=DBG%20MT5&graph_type=galaxy")

    assert page.status_code == 200
    assert "function galaxyLocatorNodes" in page.text
    assert "function galaxyLocatorLayout" in page.text
    assert "function galaxyLocatorStatusColor" in page.text
    assert "function galaxyDispatchLocatorClick" in page.text
    assert "locatorCanvas.addEventListener('click',galaxyDispatchLocatorClick,true)" in page.text

    nodes_helper = page.text.split("function galaxyLocatorNodes", 1)[1].split(
        "function galaxyLocatorLayout", 1
    )[0]
    assert "accounts()" in nodes_helper
    assert "expandedRelationGroups" not in nodes_helper
    assert "graphNodes()" not in nodes_helper

    palette = page.text.split("function galaxyLocatorStatusColor", 1)[1].split(
        "function galaxyLocatorLayout", 1
    )[0]
    assert "B:'#64748b'" in palette
    assert "M:'#f59e0b'" in palette
    assert "P:'#f97316'" in palette
    assert "T:'#ef4444'" in palette
    assert "A:'#dc2626'" in palette
    assert "TA:'#991b1b'" in palette

    dispatcher = page.text.split("function galaxyDispatchLocatorClick", 1)[1].split(
        "locatorCanvas.addEventListener", 1
    )[0]
    assert "selectedId=hit.id" in dispatcher
    assert "selectedEdgeKey=''" in dispatcher
    assert "selectedEdgeNodes=new Set()" in dispatcher
    assert "renderOverview()" in dispatcher
    assert "renderDetail()" in dispatcher


def test_kuzu_risk_defaults_to_the_current_focus_workspace(tmp_path: Path) -> None:
    app = create_account_app(make_test_settings(tmp_path))

    with TestClient(app) as client:
        page = client.get("/kuzu-risk?account=302360&platform=MT5&server=DBG%20MT5")
        unknown = client.get("/kuzu-risk?account=302360&graph_type=stale-bookmark")

    assert page.status_code == 200
    assert unknown.status_code == 200
    assert 'data-graph-type="focus-force"' in page.text
    assert 'data-graph-type="focus-force"' in unknown.text
    assert "中心约束关系工作区" in page.text
    assert "presentationGraph" in page.text
    assert "databaseStatus||n.status||'B'" in page.text
    assert "p.inProgress" in page.text
    assert "function waitForNextPoll" in page.text
    assert "state==='busy'?3000:2000" in page.text
    assert "document.hidden" in page.text
    assert "for(;;)" in page.text
    assert "poll<120" not in page.text
    assert "elapsedSeconds" in page.text
    assert "graph_type','galaxy'" in page.text
    assert "expandedGroups:new Set()" in page.text
    assert "可以直接点击图中虚线圈的边缘" in page.text
    assert "cluster-hit" in page.text
    assert "addEventListener('pointerdown',activateGroup)" in page.text
    assert "signatureOf(p)" in page.text
    assert "signature!==S.graphSignature" in page.text
    assert "const origin=S.by.get(S.subject)" in page.text
    assert "function layoutCollapsedGroups" in page.text
    assert "group.layoutX" in page.text
    assert "Math.ceil(list.length/14)" in page.text
    assert "marker-end','url(#edge-arrow)'" in page.text
    assert "scheduleRender()" in page.text
    assert "最高状态" in page.text
    assert "Math.max(.05" in page.text


def test_kuzu_risk_has_an_additive_3d_preview_mode(tmp_path: Path) -> None:
    app = create_account_app(make_test_settings(tmp_path))

    with TestClient(app) as client:
        page = client.get("/kuzu-risk?account=302360&platform=MT5&server=DBG%20MT5&graph_type=focus-3d")
        chooser = client.get("/kuzu-risk?graph_type=choose")

    assert page.status_code == 200
    assert 'data-graph-type="focus-3d"' in page.text
    assert 'canvas id="scene"' in page.text
    assert "presentationGraph" in page.text
    assert "requestAnimationFrame(render)" in page.text
    assert "拖动：旋转空间" in page.text
    assert "go('focus-force')" in page.text
    assert 'data-type="focus-3d"' in chooser.text


def test_kuzu_demo_reads_a_persisted_local_evidence_graph(tmp_path: Path) -> None:
    graph_path = tmp_path / "relationship_graph_demo.kuzu"
    database = kuzu.Database(str(graph_path))
    connection = kuzu.Connection(database)
    connection.execute(
        "CREATE NODE TABLE Entity("
        "id STRING, kind STRING, label STRING, platform STRING, server STRING, "
        "detail STRING, subject BOOL, PRIMARY KEY(id))"
    )
    connection.execute(
        "CREATE REL TABLE Evidence("
        "FROM Entity TO Entity, id STRING, kind STRING, label STRING, evidence STRING)"
    )
    connection.execute(
        "CREATE (:Entity {id: 'account:2013674', kind: 'account', label: '2013674', "
        "platform: 'MT5', server: 'DBG CN MT5', detail: '', subject: true})"
    )
    connection.execute(
        "CREATE (:Entity {id: 'ea:demo', kind: 'ea_feature', label: 'Demo EA', "
        "platform: 'MT5', server: 'DBG CN MT5', detail: 'EA / 路由特征', subject: false})"
    )
    connection.execute(
        "CREATE (:Entity {id: 'account:2013675', kind: 'account', label: '2013675', "
        "platform: 'MT5', server: 'DBG CN MT5', detail: '净盈亏：12 USD', subject: false})"
    )
    connection.execute(
        "MATCH (subject:Entity), (feature:Entity) WHERE subject.id = 'account:2013674' AND feature.id = 'ea:demo' "
        "CREATE (subject)-[:Evidence {id: 'subject-ea', kind: 'ea_feature', label: 'EA / 路由特征', "
        "evidence: '[\"匹配规则：Comment + ExpertID\"]'}]->(feature)"
    )
    connection.execute(
        "MATCH (feature:Entity), (peer:Entity) WHERE feature.id = 'ea:demo' AND peer.id = 'account:2013675' "
        "CREATE (feature)-[:Evidence {id: 'ea-peer', kind: 'ea_feature', label: 'EA 特征匹配', "
        "evidence: '[\"Comment matched\"]'}]->(peer)"
    )
    connection.close()
    del connection
    del database
    gc.collect()

    settings = replace(make_test_settings(tmp_path), kuzu_demo_path=graph_path)
    app = create_account_app(settings)
    with TestClient(app) as client:
        page = client.get('/kuzu-demo')
        graph = client.get('/api/kuzu-demo/graph?depth=2')
        invalid_depth = client.get('/api/kuzu-demo/graph?depth=4')

    assert page.status_code == 200
    assert 'Kuzu 关系图试用' in page.text
    assert graph.status_code == 200
    payload = graph.json()
    assert payload['source'] == 'kuzu-local-cache'
    assert payload['depth'] == 2
    assert payload['summary'] == {'entityCount': 3, 'relationshipCount': 2}
    assert {entity['label'] for entity in payload['entities']} == {'2013674', 'Demo EA', '2013675'}
    assert payload['relationships'][0]['evidence']
    assert invalid_depth.status_code == 422


def test_kuzu_risk_graph_propagates_until_its_score_threshold(tmp_path: Path) -> None:
    graph_path = tmp_path / "relationship_risk_graph.kuzu"
    database = kuzu.Database(str(graph_path))
    connection = kuzu.Connection(database)
    connection.execute(
        "CREATE NODE TABLE Entity("
        "id STRING, kind STRING, label STRING, platform STRING, server STRING, "
        "detail STRING, subject BOOL, PRIMARY KEY(id))"
    )
    connection.execute(
        "CREATE REL TABLE Evidence("
        "FROM Entity TO Entity, id STRING, kind STRING, label STRING, evidence STRING)"
    )
    for account, subject in (("639549", "true"), ("639550", "false"), ("639551", "false"), ("639552", "false")):
        connection.execute(
            "CREATE (:Entity {id: 'account:" + account + "', kind: 'account', label: '" + account + "', "
            "platform: 'MT5', server: 'AC CN', detail: '', subject: " + subject + "})"
        )
    for edge_id, source, target, kind in (
        ("ip-1", "639549", "639550", "login_ip"),
        ("toxic-1", "639550", "639551", "toxic_sync_same"),
        ("name-1", "639551", "639552", "same_name"),
    ):
        connection.execute(
            "MATCH (source:Entity), (target:Entity) WHERE source.id = 'account:" + source
            + "' AND target.id = 'account:" + target + "' CREATE (source)-[:Evidence {id: '" + edge_id
            + "', kind: '" + kind + "', label: '" + kind + "', evidence: '[]'}]->(target)"
        )
    connection.close()
    del connection
    del database
    gc.collect()

    settings = replace(make_test_settings(tmp_path), kuzu_risk_path=graph_path)
    app = create_account_app(settings)
    with TestClient(app) as client:
        page = client.get("/kuzu-risk?graph_type=galaxy")
        graph = client.get("/api/kuzu-risk/graph?threshold=30")
        invalid_threshold = client.get("/api/kuzu-risk/graph?threshold=0")

    assert page.status_code == 200
    assert "Kuzu 关联风险扩散" in page.text
    assert "#ff2638" in page.text
    assert graph.status_code == 200
    payload = graph.json()
    assert payload["source"] == "kuzu-local-cache"
    assert payload["threshold"] == 30
    assert payload["summary"]["entityCount"] == 4
    assert next(node for node in payload["entities"] if node["label"] == "639552")["expandable"] is False
    assert invalid_threshold.status_code == 422


def test_vue_workbench_forces_account_links_through_the_server() -> None:
    root = Path(__file__).resolve().parents[1]
    main_source = (root / "frontend" / "src" / "main.ts").read_text(encoding="utf-8")
    workbench_source = (root / "frontend" / "src" / "pages" / "WorkbenchPage.vue").read_text(encoding="utf-8")

    assert "AccountPage" not in main_source
    assert "window.location.assign(to.fullPath)" in main_source
    assert "router.push" not in workbench_source
    assert "<RouterLink" not in workbench_source
    assert "window.location.assign(accountHref(" in workbench_source
    assert "lookupMatches" in workbench_source
    assert "matches.length > 1" in workbench_source
    assert "选择平台 / 服务器" in workbench_source


def test_vue_index_is_not_cached_across_production_deployments(tmp_path: Path) -> None:
    settings = make_test_settings(tmp_path)
    settings.frontend_dist.mkdir(parents=True)
    (settings.frontend_dist / "index.html").write_text("<html><body>versioned-ui</body></html>", encoding="utf-8")
    app = create_account_app(settings)

    with TestClient(app) as client:
        response = client.get("/")

    assert response.status_code == 200
    assert "no-store" in response.headers["cache-control"]
    assert response.headers["pragma"] == "no-cache"


def test_rebate_account_audit_is_exposed_on_the_main_service(tmp_path: Path, monkeypatch) -> None:
    calls = []

    def fake_call(_self, name, *args):
        calls.append((name, args))
        return {"ok": True, "account": {"account": 7798437}, "query": {"start": "2021-06-02 00:00:00"}}

    monkeypatch.setattr(LegacyBridge, "call", fake_call)
    app = create_account_app(make_test_settings(tmp_path))
    with TestClient(app) as client:
        response = client.get("/api/rebate-churning/accounts/7798437?environment=dbg_cn")

    assert response.status_code == 200
    assert response.json()["account"]["account"] == 7798437
    assert calls == [("rebate_churning_account_audit_payload", ("7798437", "", "", "dbg_cn", ""))]


def test_rebate_account_audit_returns_account_candidates(tmp_path: Path, monkeypatch) -> None:
    class AmbiguousAccount(ValueError):
        candidates = [{"environment": "gb", "serverCode": "1"}]

    def fake_call(_self, _name, *_args):
        raise AmbiguousAccount("该账户在多个服务器存在")

    monkeypatch.setattr(LegacyBridge, "call", fake_call)
    app = create_account_app(make_test_settings(tmp_path))
    with TestClient(app) as client:
        response = client.get("/api/rebate-churning/accounts/5000000")

    assert response.status_code == 409
    assert response.json()["candidates"][0]["serverCode"] == "1"


def test_rebate_scan_submission_is_persistent_and_normalized(tmp_path: Path) -> None:
    app = create_account_app(make_test_settings(tmp_path))
    with TestClient(app) as client:
        response = client.post("/api/rebate-churning/scans", json={
            "start": "2026-07-13", "end": "2026-07-20", "environments": ["dbg_vn", "gb"],
        })
        assert response.status_code == 200
        job = response.json()["job"]
        assert job["kind"] == "rebate_churning_scan"
        assert job["payload"]["environments"] == ["dbg_vn", "gb"]
        polled = client.get(f"/api/rebate-churning/scans/{job['id']}").json()
        assert polled["status"] == "queued"


def test_rebate_scan_rejects_ranges_over_31_days(tmp_path: Path) -> None:
    app = create_account_app(make_test_settings(tmp_path))
    with TestClient(app) as client:
        response = client.post("/api/rebate-churning/scans", json={"start": "2026-06-01", "end": "2026-07-20"})
    assert response.status_code == 400


def test_ea_comment_profit_is_exposed_on_the_main_service(tmp_path: Path, monkeypatch) -> None:
    calls = []

    def fake_call(_self, name, *args):
        calls.append((name, args))
        return {"ok": True, "account": "7798437", "detected": True, "groups": [{"comment": "GoldBot"}]}

    monkeypatch.setattr(LegacyBridge, "call", fake_call)
    app = create_account_app(make_test_settings(tmp_path))
    with TestClient(app) as client:
        response = client.get("/api/accounts/by-login/7798437/ea-comment-profit?platform=MT4&server=DBG%20MT4%20CN1")

    assert response.status_code == 200
    assert response.json()["groups"][0]["comment"] == "GoldBot"
    assert calls == [("account_ea_comment_profit_payload", ("7798437", {
        "platform": "MT4", "server": "DBG MT4 CN1", "symbol": "", "start": "", "end": "",
    }))]


def test_copy_and_ea_profit_reports_download_from_the_main_service(tmp_path: Path, monkeypatch) -> None:
    calls = []

    def fake_call(_self, name, *args):
        calls.append((name, args))
        if name == "account_copy_origins_payload":
            return {"ok": True, "account": "7798437", "detected": False, "origins": [], "errors": []}
        if name == "account_ea_comment_profit_payload":
            return {"ok": True, "account": "7798437", "detected": False, "groups": [], "errors": []}
        raise AssertionError(name)

    monkeypatch.setattr(LegacyBridge, "call", fake_call)
    app = create_account_app(make_test_settings(tmp_path))
    query = "platform=MT4&server=DBG%20MT4%20CN1"
    with TestClient(app) as client:
        copy_response = client.get(f"/api/accounts/by-login/7798437/copy-report.xlsx?{query}")
        ea_response = client.get(f"/api/accounts/by-login/7798437/ea-report.xlsx?{query}")

    assert copy_response.status_code == 200
    assert copy_response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert 'filename="copy_profit_7798437_' in copy_response.headers["content-disposition"]
    assert load_workbook(BytesIO(copy_response.content), read_only=True).sheetnames == ["单主汇总"]
    assert ea_response.status_code == 200
    assert 'filename="ea_profit_7798437_' in ea_response.headers["content-disposition"]
    assert load_workbook(BytesIO(ea_response.content), read_only=True).sheetnames == ["EA汇总", "EA明细"]
    filters = {"platform": "MT4", "server": "DBG MT4 CN1", "symbol": "", "start": "", "end": ""}
    assert sorted(name for name, _args in calls) == [
        "account_copy_origins_payload",
        "account_ea_comment_profit_payload",
    ]
    assert all(args == ("7798437", filters) for _name, args in calls)


def test_push_discovery_filters_are_validated_and_persisted(tmp_path: Path) -> None:
    app = create_account_app(make_test_settings(tmp_path))
    with TestClient(app) as client:
        response = client.post("/api/push-discovery/start", json={
            "days": 3,
            "deepLimit": 80,
            "requirePeriodProfit": False,
            "limitOrders": True,
            "maxOrders": 100,
            "requireMaxLot": True,
            "minMaxLot": 0.05,
            "requireTotalProfit": True,
            "limitDeposit": True,
            "maxDeposit": 2000,
            "limitActiveRatio": True,
            "maxActiveRatio": 25,
            "excludeHandled": False,
        })
        assert response.status_code == 200
        job = response.json()["job"]
        payload = job["payload"]
        assert payload["requirePeriodProfit"] is False
        assert payload["deepLimit"] == 80
        assert payload["maxOrders"] == 100
        assert payload["requireMaxLot"] is True
        assert payload["minMaxLot"] == 0.05
        assert payload["requireTotalProfit"] is True
        assert payload["maxDeposit"] == 2000
        assert payload["maxActiveRatio"] == 25
        assert payload["excludeHandled"] is False
        assert job["max_attempts"] == 2


def test_push_discovery_active_job_can_be_resumed(tmp_path: Path) -> None:
    app = create_account_app(make_test_settings(tmp_path))
    with TestClient(app) as client:
        created = client.post("/api/push-discovery/start", json={"days": 3}).json()["job"]
        response = client.get("/api/push-discovery/active")

    assert response.status_code == 200
    assert response.json()["job"]["id"] == created["id"]
    assert response.json()["job"]["status"] == "queued"


def test_push_discovery_rejects_invalid_active_ratio(tmp_path: Path) -> None:
    app = create_account_app(make_test_settings(tmp_path))
    with TestClient(app) as client:
        response = client.post("/api/push-discovery/start", json={"maxActiveRatio": 101})
        assert response.status_code == 400


def test_push_discovery_rejects_invalid_deep_limit(tmp_path: Path) -> None:
    app = create_account_app(make_test_settings(tmp_path))
    with TestClient(app) as client:
        response = client.post("/api/push-discovery/start", json={"deepLimit": 301})
        assert response.status_code == 400


def test_push_discovery_rejects_invalid_min_max_lot(tmp_path: Path) -> None:
    app = create_account_app(make_test_settings(tmp_path))
    with TestClient(app) as client:
        response = client.post("/api/push-discovery/start", json={"minMaxLot": -0.01})
        assert response.status_code == 400


def test_bonus_arbitrage_scan_options_are_validated_and_persisted(tmp_path: Path) -> None:
    app = create_account_app(make_test_settings(tmp_path))
    with TestClient(app) as client:
        response = client.post("/api/bonus-arbitrage/scans", json={
            "start": "2026-07-01 00:00:00",
            "end": "2026-07-20 00:00:00",
            "environments": ["ac_gb", "dbg_cn"],
            "deepLimit": 80,
            "minGrant": 100,
            "excludeHandled": False,
        })

    assert response.status_code == 200
    payload = response.json()["job"]["payload"]
    assert payload["environments"] == ["ac_gb", "dbg_cn"]
    assert payload["deepLimit"] == 80
    assert payload["minGrant"] == 100
    assert payload["excludeHandled"] is False


def test_bonus_arbitrage_scan_can_be_resumed_and_rejects_invalid_range(tmp_path: Path) -> None:
    app = create_account_app(make_test_settings(tmp_path))
    with TestClient(app) as client:
        created = client.post("/api/bonus-arbitrage/scans", json={
            "start": "2026-07-01", "end": "2026-07-20",
        }).json()["job"]
        active = client.get("/api/bonus-arbitrage/scans/active")
        invalid = client.post("/api/bonus-arbitrage/scans", json={
            "start": "2025-01-01", "end": "2026-07-20",
        })

    assert active.status_code == 200
    assert active.json()["job"]["id"] == created["id"]
    assert invalid.status_code == 400
    assert "180天" in invalid.json()["error"]


def test_position_risk_scan_is_persistent_recoverable_and_bounded(tmp_path: Path) -> None:
    app = create_account_app(make_test_settings(tmp_path))
    with TestClient(app) as client:
        created = client.post("/api/position-risk/scans", json={
            "start": "2026-07-01", "end": "2026-07-22", "environments": ["ac_gb", "dbg_cn"],
            "deepLimit": 80, "excludeHandled": False, "minPositionPercent": 30,
            "minLots": 2.5, "minProfit": 100,
        })
        assert created.status_code == 200
        job = created.json()["job"]
        active = client.get("/api/position-risk/scans/active")
        polled = client.get(f"/api/position-risk/scans/{job['id']}")
        invalid = client.post("/api/position-risk/scans", json={"start": "2026-01-01", "end": "2026-07-22"})

    assert job["kind"] == "position_risk_scan"
    assert job["payload"]["deepLimit"] == 80
    assert job["payload"]["excludeHandled"] is False
    assert job["payload"]["minPositionPercent"] == 30
    assert job["payload"]["minLots"] == 2.5
    assert job["payload"]["minProfit"] == 100
    assert active.json()["job"]["id"] == job["id"]
    assert polled.json()["status"] == "queued"
    assert invalid.status_code == 400
    assert "90天" in invalid.json()["error"]


def test_job_polling_keeps_legacy_and_v2_contracts(tmp_path: Path) -> None:
    app = create_account_app(make_test_settings(tmp_path))
    with TestClient(app) as client:
        created = app.state.database.create_job("toxic_check", {"account": "302360"})
        app.state.database.update_job(
            created["id"],
            status="done",
            progress=100,
            result={
                "status": "done",
                "percent": 100,
                "message": "Toxic 检测完成",
                "result": {"account": "302360", "results": [{"type": "market_pushing", "score": 62.1}]},
            },
        )
        payload = client.get(f"/api/toxic/jobs/{created['id']}").json()
        assert payload["status"] == "done"
        assert payload["progress"] == 100
        assert payload["job"]["percent"] == 100
        assert payload["job"]["message"] == "Toxic 检测完成"
        assert payload["job"]["result"]["results"][0]["type"] == "market_pushing"


def test_queued_toxic_job_has_a_visible_legacy_status(tmp_path: Path) -> None:
    app = create_account_app(make_test_settings(tmp_path))
    with TestClient(app) as client:
        created = app.state.database.create_job("toxic_check", {"account": "239453"})
        payload = client.get(f"/api/toxic/jobs/{created['id']}").json()
        assert payload["job"]["status"] == "queued"
        assert payload["job"]["percent"] == 0
        assert payload["job"]["message"] == "已提交，等待后台执行"
