from __future__ import annotations

import copy
from io import BytesIO

from openpyxl import load_workbook

from kdesk.infrastructure.automation_reports import build_copy_profit_report, build_ea_profit_report


def copy_payload() -> dict:
    return {
        "ok": True,
        "account": "700002",
        "detected": True,
        "refreshedAt": "2026-07-21 10:00:00",
        "origins": [{
            "account": "900001",
            "platform": "MT5",
            "server": "DBG CN MT5",
            "followers": [{
                "account": "700002",
                "platform": "MT5",
                "server": "DBG CN MT5",
                "matchedSourceOrders": 2,
                "orders": 3,
                "volume": 1.25,
                "grossProfit": 125.5,
                "commission": -2.5,
                "swap": -1,
                "taxes": 0,
                "netProfit": 122,
                "displayCurrency": "USD",
                "isCentAccount": False,
                "isCurrentAccount": True,
                "symbols": ["XAUUSD"],
                "tickets": ["100001", "100002"],
                "firstTime": "2026-07-01 10:00:00",
                "lastTime": "2026-07-02 11:00:00",
            }],
            "followerSummary": {
                "accounts": 1,
                "profitableAccounts": 1,
                "losingAccounts": 0,
                "orders": 3,
                "volume": 1.25,
                "grossProfit": 125.5,
                "commission": -2.5,
                "swap": -1,
                "taxes": 0,
                "netProfit": 122,
                "currency": "USD",
            },
            "followerDiscovery": {},
            "followerOrders": [{
                "account": "700002",
                "platform": "MT5",
                "server": "DBG CN MT5",
                "ticket": "100001",
                "matchedSourceOrderIds": ["80001"],
                "symbol": "XAUUSD",
                "openTime": "2026-07-01 10:00:00",
                "closeTime": "2026-07-01 10:05:00",
                "volume": 1.25,
                "grossProfit": 125.5,
                "commission": -2.5,
                "swap": -1,
                "taxes": 0,
                "netProfit": 122,
                "displayCurrency": "USD",
            }],
            "sourceOrders": [{
                "orderId": "80001", "ticket": "90001", "symbol": "XAUUSD", "time": "2026-07-01 10:00:00",
            }],
        }],
        "errors": [],
    }


def signal_payload() -> dict:
    return {
        "ok": True,
        "account": "700002",
        "detected": True,
        "definition": "同一 Signal 标识的账户视为同组。",
        "refreshedAt": "2026-07-21 10:00:01",
        "groups": [{
            "signalTag": "Signal #ABC IN",
            "platform": "MT4",
            "server": "DBG MT4 CN1",
            "database": "DBG",
            "members": [{
                "account": "700002",
                "database": "DBG",
                "platform": "MT4",
                "server": "DBG MT4 CN1",
                "status": "Enabled",
                "closedOrders": 4,
                "openOrders": 1,
                "closedLots": 0.8,
                "closedNetProfit": -20,
                "floatingNetProfit": 5,
                "combinedNetProfit": -15,
                "rebate": 3.5,
                "currency": "USD",
                "isCentAccount": False,
                "firstClose": "2026-07-01 10:00:00",
                "lastClose": "2026-07-03 10:00:00",
            }],
            "totals": {
                "accounts": 1,
                "profitableAccounts": 0,
                "losingAccounts": 1,
                "closedOrders": 4,
                "openOrders": 1,
                "closedLots": 0.8,
                "closedNetProfit": -20,
                "floatingNetProfit": 5,
                "combinedNetProfit": -15,
                "rebate": 3.5,
            },
            "limitations": [],
        }],
        "errors": [],
    }


def ea_payload() -> dict:
    return {
        "ok": True,
        "account": "700002",
        "detected": True,
        "definition": "相同 EA comment 的账户归为一组。",
        "refreshedAt": "2026-07-21 10:05:00",
        "groups": [{
            "comment": "GoldBot",
            "classification": "exact_ea",
            "countedAsEa": True,
            "expertId": 42,
            "matchRule": "同服务器要求 Comment 与 ExpertID/MAGIC 同时相同；跨服务器按 Comment 相同匹配。",
            "database": "DBG",
            "platform": "MT4",
            "server": "DBG MT4 CN1",
            "currentOrders": 2,
            "currentVolume": 0.4,
            "currentNetProfit": 18,
            "firstTime": "2026-07-01 10:00:00",
            "lastTime": "2026-07-02 10:00:00",
            "members": [{
                "account": "700002",
                "database": "DBG",
                "platform": "MT4",
                "server": "DBG MT4 CN1",
                "orders": 2,
                "volume": 0.4,
                "grossProfit": 20,
                "commission": -1,
                "swap": -1,
                "taxes": 0,
                "netProfit": 18,
                "currency": "USD",
                "isCentAccount": False,
                "isCurrentAccount": True,
                "expertIds": ["42"],
                "matchClues": ["同服务器：Comment「GoldBot」相同，MAGIC 42 相同"],
                "symbols": ["XAUUSD"],
                "tickets": ["200001", "200002"],
                "firstTime": "2026-07-01 10:00:00",
                "lastTime": "2026-07-02 10:00:00",
            }],
            "totals": {
                "accounts": 1,
                "profitableAccounts": 1,
                "losingAccounts": 0,
                "flatAccounts": 0,
                "orders": 2,
                "volume": 0.4,
                "grossProfit": 20,
                "commission": -1,
                "swap": -1,
                "taxes": 0,
                "netProfit": 18,
                "currency": "USD",
            },
            "limitations": [],
        }],
        "errors": [],
    }


def test_copy_report_is_organized_by_master_with_summary_and_order_details() -> None:
    content = build_copy_profit_report(
        copy_payload(), signal_payload(), {"platform": "MT4", "server": "DBG MT4 CN1"},
        exported_at="2026-07-21 10:10:00",
    )
    workbook = load_workbook(BytesIO(content), data_only=False)

    assert workbook.sheetnames == ["单主汇总", "单主_900001"]
    assert workbook["单主汇总"]["A1"].value == "单主跟单总盈亏汇总"
    assert workbook["单主汇总"]["A4"].value == "900001"
    assert workbook["单主汇总"]["G4"].value == 122
    master_sheet = workbook["单主_900001"]
    assert master_sheet["A1"].value == "单主 900001 跟单收益"
    assert master_sheet["A8"].value == "700002"
    assert master_sheet["K8"].value == 122
    assert master_sheet["K8"].number_format == '#,##0.00;[Red]-#,##0.00;0.00'
    assert master_sheet.tables["CopyMasterFollowers1"].ref == "A7:O8"
    assert master_sheet["A12"].value == "700002"
    assert master_sheet["D12"].value == "100001"
    assert master_sheet["E12"].value == "80001"
    assert master_sheet.tables["CopyMasterOrders1"].ref == "A11:O12"


def test_ea_report_preserves_account_ids_and_profit_reconciliation() -> None:
    content = build_ea_profit_report(
        ea_payload(), {"platform": "MT4", "server": "DBG MT4 CN1"}, exported_at="2026-07-21 10:10:00"
    )
    workbook = load_workbook(BytesIO(content), data_only=False)

    assert workbook.sheetnames == ["EA汇总", "EA账户明细", "导出说明"]
    assert workbook["EA汇总"]["A1"].value == "EA Comment 收益汇总 - 账号 700002"
    detail_sheet = workbook["EA账户明细"]
    assert detail_sheet["G4"].value == "700002"
    assert detail_sheet["G4"].data_type == "s"
    assert detail_sheet["B4"].value == "42"
    assert "MAGIC 42 相同" in detail_sheet["C4"].value
    assert detail_sheet["O4"].value == 18
    assert detail_sheet["P4"].value == "=ROUND(O4-SUM(K4:N4),2)"
    assert detail_sheet.tables["EaAccountDetails"].ref == "A3:V4"
    assert workbook["导出说明"]["B5"].value == "净盈亏 = 毛盈亏 + 手续费/Fee + 利息/Swap + 税费。"
    assert "Signal" not in workbook["导出说明"]["B5"].value


def test_ea_report_excludes_possible_copy_routes_from_headline_kpis_but_keeps_detail() -> None:
    payload = ea_payload()
    route = copy.deepcopy(payload["groups"][0])
    route.update({
        "comment": "1/521/{SOURCE_ID}",
        "classification": "possible_copy_route",
        "classificationLabel": "可能是跟单路由",
        "countedAsEa": False,
    })
    payload["groups"].append(route)

    workbook = load_workbook(BytesIO(build_ea_profit_report(payload, {})), data_only=False)

    summary = workbook["EA汇总"]
    assert summary["A7"].value == 1
    assert summary["D7"].value == 1
    assert summary["A12"].value == "[可能是跟单路由] 1/521/{SOURCE_ID}"
    detail_values = [workbook["EA账户明细"].cell(row, 1).value for row in range(4, 6)]
    assert "[可能是跟单路由] 1/521/{SOURCE_ID}" in detail_values
