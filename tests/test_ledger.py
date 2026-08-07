from __future__ import annotations

from pathlib import Path
from threading import Barrier, Thread

from openpyxl import load_workbook
from sqlalchemy import text

from kdesk.application.ledger_service import LedgerService
from kdesk.infrastructure.database import Database


def test_sqlite_ledger_keeps_history_and_round_trips_excel(tmp_path: Path) -> None:
    database = Database(tmp_path / "kdesk.sqlite")
    database.create_schema()
    service = LedgerService(database)

    created = service.save({"账号": "302360", "建议动作": "M", "状态": "待复核", "风险标签": "EA"})
    record_id = created["record"]["记录ID"]
    updated = service.save({"状态": "观察中", "风险/问题备注": "review"}, record_id)

    assert updated["record"]["账号"] == "302360"
    assert updated["record"]["状态"] == "观察中"
    assert len(database.history(record_id)) == 2

    target = service.export_excel(tmp_path / "export.xlsx")
    workbook = load_workbook(target, read_only=True, data_only=True)
    assert "问题账户" in workbook.sheetnames
    assert "修改历史" in workbook.sheetnames
    assert workbook["问题账户"].max_row == 2
    workbook.close()


def test_persistent_job_survives_database_reopen(tmp_path: Path) -> None:
    path = tmp_path / "jobs.sqlite"
    first = Database(path)
    first.create_schema()
    created = first.create_job("kline_inspect", {"statement": "sample.html"})

    second = Database(path)
    second.create_schema()
    claimed = second.claim_next_job()

    assert claimed is not None
    assert claimed["id"] == created["id"]
    assert claimed["status"] == "running"

    recovered = second.recover_interrupted_jobs()
    assert recovered == 1
    assert second.get_job(created["id"])["status"] == "queued"


def test_queued_job_can_be_cancelled(tmp_path: Path) -> None:
    database = Database(tmp_path / "jobs.sqlite")
    database.create_schema()
    created = database.create_job("kline_inspect", {"statement": "sample.html"})
    cancelled = database.request_job_cancel(created["id"])
    assert cancelled is not None
    assert cancelled["status"] == "cancelled"
    assert database.claim_next_job() is None


def test_job_idempotency_only_deduplicates_active_runs(tmp_path: Path) -> None:
    database = Database(tmp_path / "jobs.sqlite")
    database.create_schema()
    first = database.create_job("push_discovery", {"days": 7}, idempotency_key="push:7")
    duplicate = database.create_job("push_discovery", {"days": 7}, idempotency_key="push:7")
    assert duplicate["id"] == first["id"]

    claimed = database.claim_next_job()
    assert claimed is not None
    database.update_job(claimed["id"], status="done", progress=100, result={"ok": True})
    rerun = database.create_job("push_discovery", {"days": 7}, idempotency_key="push:7")
    assert rerun["id"] != first["id"]


def test_active_job_prefers_running_over_queued(tmp_path: Path) -> None:
    database = Database(tmp_path / "jobs.sqlite")
    database.create_schema()
    running = database.create_job("push_discovery", {"days": 3})
    database.claim_next_job(kinds=("push_discovery",))
    database.create_job("push_discovery", {"days": 7})

    active = database.get_active_job("push_discovery")

    assert active is not None
    assert active["id"] == running["id"]
    assert active["status"] == "running"


def test_job_queues_claim_only_their_assigned_kinds(tmp_path: Path) -> None:
    database = Database(tmp_path / "jobs.sqlite")
    database.create_schema()
    discovery = database.create_job("push_discovery", {"days": 7})
    toxic = database.create_job("toxic_check", {"account": "239453"})

    interactive_claim = database.claim_next_job(kinds=("toxic_check", "kline_from_database"))
    discovery_claim = database.claim_next_job(kinds=("push_discovery",))

    assert interactive_claim is not None
    assert interactive_claim["id"] == toxic["id"]
    assert discovery_claim is not None
    assert discovery_claim["id"] == discovery["id"]


def test_worker_recovery_is_scoped_to_its_queue(tmp_path: Path) -> None:
    database = Database(tmp_path / "jobs.sqlite")
    database.create_schema()
    discovery = database.create_job("push_discovery", {"days": 7})
    toxic = database.create_job("toxic_check", {"account": "239453"})
    database.claim_next_job(kinds=("push_discovery",))
    database.claim_next_job(kinds=("toxic_check",))

    assert database.recover_interrupted_jobs(kinds=("toxic_check",)) == 1
    assert database.get_job(toxic["id"])["status"] == "queued"
    assert database.get_job(discovery["id"])["status"] == "running"


def test_worker_recovery_does_not_reset_a_fresh_lease(tmp_path: Path) -> None:
    database = Database(tmp_path / "jobs.sqlite")
    database.create_schema()
    created = database.create_job("push_discovery", {"days": 7})
    claimed = database.claim_next_job(kinds=("push_discovery",))
    assert claimed is not None

    assert database.recover_interrupted_jobs(kinds=("push_discovery",), stale_after_seconds=180) == 0
    assert database.get_job(created["id"])["status"] == "running"

    with database.session() as session:
        session.execute(
            text("update job_runs set heartbeat_at = '2020-01-01 00:00:00' where job_id = :job_id"),
            {"job_id": created["id"]},
        )
    assert database.recover_interrupted_jobs(kinds=("push_discovery",), stale_after_seconds=180) == 1
    assert database.get_job(created["id"])["status"] == "queued"


def test_touch_job_refreshes_only_running_jobs(tmp_path: Path) -> None:
    database = Database(tmp_path / "jobs.sqlite")
    database.create_schema()
    created = database.create_job("push_discovery", {"days": 7})
    database.touch_job(created["id"])
    assert database.get_job(created["id"])["heartbeat_at"] == ""
    database.claim_next_job(kinds=("push_discovery",))
    database.touch_job(created["id"])
    assert database.get_job(created["id"])["heartbeat_at"]


def test_concurrent_workers_claim_a_job_once(tmp_path: Path) -> None:
    database = Database(tmp_path / "jobs.sqlite")
    database.create_schema()
    created = database.create_job("push_discovery", {"days": 7})
    barrier = Barrier(2)
    claims: list[dict | None] = []

    def claim() -> None:
        barrier.wait()
        claims.append(database.claim_next_job(kinds=("push_discovery",)))

    threads = [Thread(target=claim) for _ in range(2)]
    for thread in threads:
        thread.start()
    for thread in threads:
        thread.join()

    assert sum(item is not None for item in claims) == 1
    assert database.get_job(created["id"])["attempts"] == 1


def test_ledger_service_batch_import_preview_and_delete(tmp_path: Path) -> None:
    source_database = Database(tmp_path / "source.sqlite")
    source_database.create_schema()
    source_service = LedgerService(source_database)
    batch = source_service.save_many({"建议动作": "P", "状态": "观察中"}, ["10001", "10002", ""])
    assert len(batch["records"]) == 2
    assert source_service.list_payload()["summary"]["total"] == 2

    workbook = source_service.export_excel(tmp_path / "source.xlsx")
    target_database = Database(tmp_path / "target.sqlite")
    target_database.create_schema()
    target_service = LedgerService(target_database)
    preview = target_service.import_preview(workbook)
    assert preview["added"] == 2
    imported = target_service.import_excel(workbook)
    assert imported["imported"] is True
    assert target_database.count_accounts() == 2

    record_id = target_database.find_by_login("10001").record_id
    assert target_service.delete(record_id)["ok"] is True
    assert target_service.delete(record_id)["ok"] is False
